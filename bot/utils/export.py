"""
Функции экспорта данных в Excel (XLSX).
"""

import io
from datetime import date
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from bot.database import get_session, crud
from bot.database.models import Subject, Student, Attendance


def create_attendance_report(subject_id: int) -> io.BytesIO:
    """
    Создать отчёт о посещаемости по дисциплине.
    Возвращает файл в памяти (BytesIO).
    """
    session = get_session()
    
    try:
        subject = crud.get_subject_by_id(session, subject_id)
        students = crud.get_students_by_subject(session, subject_id)
        dates = crud.get_subject_attendance_dates(session, subject_id)
        
        # Создаём книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Посещаемость"
        
        # Стили
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        
        present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок таблицы
        ws['A1'] = f"Посещаемость: {subject.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, len(dates) + 3))
        
        # Заголовки колонок
        row = 3
        ws.cell(row=row, column=1, value="№").font = header_font_white
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value="ФИО").font = header_font_white
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=2).border = thin_border
        
        # Колонки дат
        col = 3
        for d in dates:
            cell = ws.cell(row=row, column=col, value=d.strftime("%d.%m"))
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            col += 1
        
        # Колонки итогов
        ws.cell(row=row, column=col, value="Всего").font = header_font_white
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).alignment = center_align
        ws.cell(row=row, column=col).border = thin_border
        
        ws.cell(row=row, column=col + 1, value="%").font = header_font_white
        ws.cell(row=row, column=col + 1).fill = header_fill
        ws.cell(row=row, column=col + 1).alignment = center_align
        ws.cell(row=row, column=col + 1).border = thin_border
        
        # Данные студентов
        row = 4
        for idx, student in enumerate(students, 1):
            # Номер
            ws.cell(row=row, column=1, value=idx).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border
            
            # ФИО
            ws.cell(row=row, column=2, value=student.full_name).border = thin_border
            
            # Посещаемость по датам
            col = 3
            present_count = 0
            
            for d in dates:
                attendance_data = crud.get_attendance_by_subject_and_date(session, subject_id, d)
                is_present = attendance_data.get(student.id, False)
                
                cell = ws.cell(row=row, column=col)
                if is_present:
                    cell.value = "+"
                    cell.fill = present_fill
                    present_count += 1
                else:
                    cell.value = "-"
                    cell.fill = absent_fill
                
                cell.alignment = center_align
                cell.border = thin_border
                col += 1
            
            # Итоги
            total_dates = len(dates)
            percent = round(present_count / total_dates * 100) if total_dates > 0 else 0
            
            ws.cell(row=row, column=col, value=present_count).alignment = center_align
            ws.cell(row=row, column=col).border = thin_border
            
            ws.cell(row=row, column=col + 1, value=f"{percent}%").alignment = center_align
            ws.cell(row=row, column=col + 1).border = thin_border
            
            row += 1
        
        # Итоговая строка
        if students and dates:
            row += 1
            ws.cell(row=row, column=1, value="").border = thin_border
            ws.cell(row=row, column=2, value="ИТОГО").font = header_font
            ws.cell(row=row, column=2).border = thin_border
            
            col = 3
            for d in dates:
                attendance_data = crud.get_attendance_by_subject_and_date(session, subject_id, d)
                present = sum(1 for s in students if attendance_data.get(s.id, False))
                
                cell = ws.cell(row=row, column=col, value=f"{present}/{len(students)}")
                cell.alignment = center_align
                cell.font = Font(bold=True)
                cell.border = thin_border
                col += 1
        
        # Автоширина колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for i in range(3, col + 2):
            ws.column_dimensions[get_column_letter(i)].width = 8
        
        # Сохраняем в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    finally:
        session.close()


def create_all_subjects_report(teacher_id: int) -> io.BytesIO:
    """
    Создать общий отчёт по всем дисциплинам преподавателя.
    Каждая дисциплина на отдельном листе.
    """
    session = get_session()
    
    try:
        subjects = crud.get_subjects_by_teacher(session, teacher_id)
        
        wb = Workbook()
        # Удаляем дефолтный лист
        wb.remove(wb.active)
        
        if not subjects:
            ws = wb.create_sheet("Нет данных")
            ws['A1'] = "Нет дисциплин"
        else:
            for subject in subjects:
                ws = wb.create_sheet(subject.name[:31])  # Имя листа до 31 символа
                _fill_subject_sheet(ws, subject, session)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    finally:
        session.close()


def _fill_subject_sheet(ws, subject: Subject, session):
    """Заполнить лист данными о посещаемости дисциплины."""
    students = crud.get_students_by_subject(session, subject.id)
    dates = crud.get_subject_attendance_dates(session, subject.id)
    
    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    row = 1
    ws.cell(row=row, column=1, value="№").font = header_font_white
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = center_align
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value="ФИО").font = header_font_white
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=2).border = thin_border
    
    col = 3
    for d in dates:
        cell = ws.cell(row=row, column=col, value=d.strftime("%d.%m"))
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        col += 1
    
    ws.cell(row=row, column=col, value="Всего").font = header_font_white
    ws.cell(row=row, column=col).fill = header_fill
    ws.cell(row=row, column=col).alignment = center_align
    ws.cell(row=row, column=col).border = thin_border
    
    ws.cell(row=row, column=col + 1, value="%").font = header_font_white
    ws.cell(row=row, column=col + 1).fill = header_fill
    ws.cell(row=row, column=col + 1).alignment = center_align
    ws.cell(row=row, column=col + 1).border = thin_border
    
    # Данные
    row = 2
    for idx, student in enumerate(students, 1):
        ws.cell(row=row, column=1, value=idx).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=student.full_name).border = thin_border
        
        col = 3
        present_count = 0
        
        for d in dates:
            attendance_data = crud.get_attendance_by_subject_and_date(session, subject.id, d)
            is_present = attendance_data.get(student.id, False)
            
            cell = ws.cell(row=row, column=col)
            if is_present:
                cell.value = "+"
                cell.fill = present_fill
                present_count += 1
            else:
                cell.value = "-"
                cell.fill = absent_fill
            
            cell.alignment = center_align
            cell.border = thin_border
            col += 1
        
        total_dates = len(dates)
        percent = round(present_count / total_dates * 100) if total_dates > 0 else 0
        
        ws.cell(row=row, column=col, value=present_count).alignment = center_align
        ws.cell(row=row, column=col).border = thin_border
        
        ws.cell(row=row, column=col + 1, value=f"{percent}%").alignment = center_align
        ws.cell(row=row, column=col + 1).border = thin_border
        
        row += 1
    
    # Автоширина
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    for i in range(3, col + 2):
        ws.column_dimensions[get_column_letter(i)].width = 8

